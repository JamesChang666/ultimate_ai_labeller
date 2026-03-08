import sys, re, os, json
from collections import defaultdict, Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

def load_data(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip().lstrip("\ufeff").lower().replace(" ", "_") for c in df.columns]

    has_golden = "status" in df.columns
    records = []
    for _, row in df.iterrows():
        name     = str(row.get("image_name", "")).strip()
        detected = str(row.get("detected_classes", "")).strip()
        classes  = parse_classes(detected)
        status   = str(row.get("status", "N/A")).strip() if has_golden else "N/A"
        details  = str(row.get("details", "")).strip() if has_golden else ""
        iou_val  = parse_iou_from_details(details)
        matched, total = parse_matched(details)
        records.append({
            "timestamp":        str(row.get("timestamp", "")).strip(),
            "image_name":       name,
            "prefix":           get_prefix(name),
            "detected_classes": detected,
            "golden_mode":      str(row.get("golden_mode", "—")).strip() if has_golden else "—",
            "iou_threshold":    str(row.get("iou_threshold", "—")).strip() if has_golden else "—",
            "status":           status,
            "details":          details,
            "avg_iou":          iou_val,
            "matched":          matched,
            "total":            total,
            "total_components": sum(classes.values()),
            "num_classes":      len(classes),
            "classes_dict":     classes,
            "has_golden":       has_golden,
        })
    return records, has_golden


def parse_classes(s):
    result = {}
    for item in s.split(";"):
        m = re.match(r"(.+)\s+x(\d+)", item.strip())
        if m:
            result[m.group(1).strip()] = int(m.group(2))
    return result


def parse_iou_from_details(d):
    m = re.search(r"avg IoU=([0-9.]+)", str(d))
    return float(m.group(1)) if m else None


def parse_matched(d):
    m = re.search(r"(\d+)/(\d+) matched", str(d))
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def get_prefix(name):
    for prefix in ("back_crop", "empty_crop", "train_crop", "crop"):
        if name.startswith(prefix):
            return prefix
    return "other"


def aggregate(records):
    class_totals    = defaultdict(int)
    class_img_count = defaultdict(int)
    prefix_stats    = defaultdict(lambda: {
        "count": 0, "total_components": 0,
        "pass": 0, "fail": 0, "iou_sum": 0, "iou_count": 0
    })
    status_counts   = Counter()
    iou_values      = []

    for r in records:
        for cls, cnt in r["classes_dict"].items():
            class_totals[cls]    += cnt
            class_img_count[cls] += 1
        ps = prefix_stats[r["prefix"]]
        ps["count"]            += 1
        ps["total_components"] += r["total_components"]
        s = r["status"].upper()
        status_counts[s] += 1
        if s == "PASS": ps["pass"] += 1
        elif s == "FAIL": ps["fail"] += 1
        if r["avg_iou"] is not None:
            ps["iou_sum"]   += r["avg_iou"]
            ps["iou_count"] += 1
            iou_values.append(r["avg_iou"])

    sorted_classes = sorted(class_totals.items(), key=lambda x: x[1], reverse=True)
    return sorted_classes, class_img_count, prefix_stats, status_counts, iou_values

def build_excel(records, sorted_classes, class_img_count, prefix_stats,
                status_counts, iou_values, has_golden, out_path):

    H_FILL  = PatternFill("solid", start_color="1A3A5C")
    H_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    S_FILL  = PatternFill("solid", start_color="2E75B6")
    S_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    T_FILL  = PatternFill("solid", start_color="4472C4")
    T_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    CELL_F  = Font(name="Calibri", size=10)
    ALT_F   = PatternFill("solid", start_color="EBF3FB")
    KPI_F   = PatternFill("solid", start_color="D6E8F7")
    FAIL_F  = PatternFill("solid", start_color="FDDEDE")
    FAIL_FT = Font(bold=True, color="C00000", name="Calibri", size=10)
    PASS_F  = PatternFill("solid", start_color="D8F5E4")
    PASS_FT = Font(bold=True, color="1A7A3C", name="Calibri", size=10)
    NA_F    = PatternFill("solid", start_color="F0F0F0")
    NA_FT   = Font(color="888888", name="Calibri", size=10)

    def thdr(ws, row, cols):
        for col, text in cols:
            c = ws.cell(row=row, column=col, value=text)
            c.font = T_FONT; c.fill = T_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")

    def shdr(ws, row, col, text, span=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = S_FONT; c.fill = S_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col+span-1)

    def status_style(ws, row, col, status):
        c = ws.cell(row=row, column=col)
        s = status.upper()
        if s == "FAIL":   c.fill, c.font = FAIL_F, FAIL_FT
        elif s == "PASS": c.fill, c.font = PASS_F, PASS_FT
        else:             c.fill, c.font = NA_F, NA_FT

    wb = Workbook()
    total_img  = len(records)
    total_comp = sum(r["total_components"] for r in records)
    fail_cnt   = status_counts.get("FAIL", 0)
    pass_cnt   = status_counts.get("PASS", 0)
    na_cnt     = total_img - fail_cnt - pass_cnt
    pass_rate  = round(pass_cnt / total_img * 100, 1) if total_img else 0

    ws = wb.active
    ws.title = "Raw Data"

    if has_golden:
        headers    = ["Timestamp", "Image Name", "Status", "Category",
                      "Detected Classes", "IoU Mode", "IoU Threshold",
                      "Matched/Total", "Avg IoU", "Details", "Components", "# Classes"]
        col_widths = [20, 22, 9, 12, 55, 10, 14, 14, 10, 30, 12, 10]
        STATUS_COL = 3
    else:
        headers    = ["Timestamp", "Image Name", "Category",
                      "Detected Classes", "Components", "# Classes"]
        col_widths = [20, 22, 12, 70, 12, 10]
        STATUS_COL = None

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    for i, r in enumerate(records, 2):
        row_fill = ALT_F if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        if has_golden:
            matched_str = f"{r['matched']}/{r['total']}" if r['matched'] is not None else "—"
            iou_str     = f"{r['avg_iou']:.3f}" if r['avg_iou'] is not None else "—"
            vals = [r["timestamp"], r["image_name"], r["status"], r["prefix"],
                    r["detected_classes"], r["golden_mode"], r["iou_threshold"],
                    matched_str, iou_str, r["details"],
                    r["total_components"], r["num_classes"]]
        else:
            vals = [r["timestamp"], r["image_name"], r["prefix"],
                    r["detected_classes"], r["total_components"], r["num_classes"]]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = CELL_F; cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col == (5 if has_golden else 4)))

        if STATUS_COL:
            status_style(ws, i, STATUS_COL, r["status"])

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")

    shdr(ws2, 1, 1, "📊  Detection Summary Report", 4)
    ws2.row_dimensions[1].height = 26

    kpis = [("Total Images", total_img), ("PASS", pass_cnt), ("FAIL", fail_cnt)]
    if has_golden:
        kpis.append(("Pass Rate", f"{pass_rate}%"))
        if iou_values:
            kpis.append(("Avg IoU", f"{round(sum(iou_values)/len(iou_values),3)}"))
    kpis.append(("Component Classes", len(sorted_classes)))

    ws2.cell(row=3, column=1, value="Metric").font = Font(bold=True, name="Calibri")
    ws2.cell(row=3, column=2, value="Value").font  = Font(bold=True, name="Calibri")
    for i, (metric, val) in enumerate(kpis, 4):
        ws2.cell(row=i, column=1, value=metric).font = CELL_F
        c = ws2.cell(row=i, column=2, value=val)
        c.font = Font(name="Calibri", size=11, bold=True); c.fill = KPI_F
        if metric == "PASS": c.font = Font(name="Calibri", size=11, bold=True, color="1A7A3C")
        if metric == "FAIL": c.font = Font(name="Calibri", size=11, bold=True, color="C00000")

    for col, w in zip(["A","B"], [28, 18]):
        ws2.column_dimensions[col].width = w

    start_row = 4 + len(kpis) + 2
    shdr(ws2, start_row, 1, "Results by Category", 5)
    if has_golden:
        thdr(ws2, start_row+1, [(1,"Category"),(2,"Images"),(3,"PASS"),(4,"FAIL"),(5,"Avg IoU")])
    else:
        thdr(ws2, start_row+1, [(1,"Category"),(2,"Images"),(3,"Total Components"),(4,"Avg Comp/Image"),(5,"")])

    for r_idx, (prefix, stats) in enumerate(sorted(prefix_stats.items()), start_row+2):
        fill = ALT_F if r_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        if has_golden:
            avg_iou = round(stats["iou_sum"]/stats["iou_count"], 3) if stats["iou_count"] else "—"
            vals = [prefix, stats["count"], stats["pass"], stats["fail"], avg_iou]
        else:
            avg_c = round(stats["total_components"]/stats["count"], 1) if stats["count"] else 0
            vals  = [prefix, stats["count"], stats["total_components"], avg_c, ""]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=r_idx, column=col, value=val)
            c.font = CELL_F; c.fill = fill
        if has_golden:
            p_c = ws2.cell(row=r_idx, column=3)
            p_c.font = Font(name="Calibri", size=10, bold=True, color="1A7A3C")
            f_c = ws2.cell(row=r_idx, column=4)
            f_c.font = Font(name="Calibri", size=10, bold=True, color="C00000")

    for col, w in zip(["C","D","E"], [10, 10, 12]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Class Analysis")
    shdr(ws3, 1, 1, "Component Class Detection Analysis", 4)
    thdr(ws3, 2, [(1,"Class"),(2,"Total Detected"),(3,"Images Found In"),(4,"Avg per Image")])
    for i, (cls, total) in enumerate(sorted_classes, 3):
        img_cnt = class_img_count[cls]
        avg     = round(total/img_cnt, 2) if img_cnt else 0
        fill    = ALT_F if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for col, val in enumerate([cls, total, img_cnt, avg], 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.font = CELL_F; c.fill = fill
    for col, w in zip(["A","B","C","D"], [22, 16, 20, 16]):
        ws3.column_dimensions[col].width = w

    chart = BarChart()
    chart.type = "col"; chart.title = "Components by Class"
    chart.style = 10; chart.width = 22; chart.height = 13
    last = 2 + len(sorted_classes)
    chart.add_data(Reference(ws3, min_col=2, min_row=2, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws3, min_col=1, min_row=3, max_row=last))
    ws3.add_chart(chart, "F2")

    if has_golden and iou_values:
        ws4 = wb.create_sheet("IoU Analysis")
        shdr(ws4, 1, 1, "IoU Score Analysis by Category", 4)
        thdr(ws4, 2, [(1,"Category"),(2,"Images"),(3,"PASS"),(4,"FAIL"),
                      (5,"Avg IoU"),(6,"Min IoU"),(7,"Max IoU")])

        cat_iou = defaultdict(list)
        for r in records:
            if r["avg_iou"] is not None:
                cat_iou[r["prefix"]].append((r["avg_iou"], r["status"]))

        for i, (cat, items) in enumerate(sorted(cat_iou.items()), 3):
            ious      = [x[0] for x in items]
            pass_n    = sum(1 for x in items if x[1].upper()=="PASS")
            fail_n    = sum(1 for x in items if x[1].upper()=="FAIL")
            fill      = ALT_F if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            vals      = [cat, len(items), pass_n, fail_n,
                         round(sum(ious)/len(ious),3), round(min(ious),3), round(max(ious),3)]
            for col, val in enumerate(vals, 1):
                c = ws4.cell(row=i, column=col, value=val)
                c.font = CELL_F; c.fill = fill

        for col, w in zip(["A","B","C","D","E","F","G"], [14,10,10,10,12,12,12]):
            ws4.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"  ✅ Excel → {out_path}")

def build_html(records, sorted_classes, class_img_count, prefix_stats,
               status_counts, iou_values, has_golden, out_path):

    total_img  = len(records)
    fail_cnt   = status_counts.get("FAIL", 0)
    pass_cnt   = status_counts.get("PASS", 0)
    pass_rate  = round(pass_cnt/total_img*100, 1) if total_img else 0
    total_comp = sum(r["total_components"] for r in records)
    avg_comp   = round(total_comp/total_img, 1) if total_img else 0
    avg_iou    = round(sum(iou_values)/len(iou_values), 3) if iou_values else None
    ts_label   = str(records[0]["timestamp"]) if records else "N/A"

    cat_data   = sorted(prefix_stats.items())
    grand      = sum(t for _, t in sorted_classes) or 1

    cls_labels = json.dumps([c for c, _ in sorted_classes])
    cls_totals = json.dumps([t for _, t in sorted_classes])
    cls_imgs   = json.dumps([class_img_count[c] for c, _ in sorted_classes])
    cat_labels = json.dumps([c for c, _ in cat_data])
    cat_counts = json.dumps([s["count"] for _, s in cat_data])
    cat_comps  = json.dumps([s["total_components"] for _, s in cat_data])
    cat_pass   = json.dumps([s["pass"] for _, s in cat_data])
    cat_fail   = json.dumps([s["fail"] for _, s in cat_data])

    iou_hist_labels, iou_hist_vals = [], []
    if has_golden and iou_values:
        bins = [(i/10, (i+1)/10) for i in range(10)]
        for lo, hi in bins:
            iou_hist_labels.append(f"{lo:.1f}-{hi:.1f}")
            iou_hist_vals.append(sum(1 for v in iou_values if lo <= v < hi))
        iou_hist_vals[-1] += sum(1 for v in iou_values if v == 1.0)

    palette = ["#3b82f6","#06b6d4","#8b5cf6","#ec4899","#f97316",
               "#10b981","#f59e0b","#60a5fa","#ef4444","#6366f1","#22c55e","#84cc16"]

    cls_rows = ""
    for i, (cls, total) in enumerate(sorted_classes):
        img_cnt = class_img_count[cls]
        avg     = round(total/img_cnt, 2) if img_cnt else 0
        pct     = round(total/grand*100, 1)
        bar_w   = round(total/(sorted_classes[0][1] or 1)*100)
        color   = palette[i % len(palette)]
        cls_rows += (f'<tr>'
                     f'<td><span class="dot" style="background:{color}"></span>{cls}</td>'
                     f'<td class="num">{total}</td><td class="num">{img_cnt}</td>'
                     f'<td class="num">{avg}</td>'
                     f'<td><span class="pct">{pct}%</span>'
                     f'<div class="bar-bg"><div class="bar-fg" style="width:{bar_w}%;background:{color}80"></div></div></td>'
                     f'</tr>\n')

    cat_rows = ""
    for prefix, stats in cat_data:
        avg = round(stats["total_components"]/stats["count"], 1) if stats["count"] else 0
        avg_iou_cat = f"{round(stats['iou_sum']/stats['iou_count'],3):.3f}" if stats["iou_count"] else "—"
        pass_badge = f'<span class="badge pass">{stats["pass"]} PASS</span>' if has_golden else ""
        fail_badge = f'<span class="badge fail">{stats["fail"]} FAIL</span>' if has_golden else ""
        iou_cell   = f'<td class="num">{avg_iou_cat}</td>' if has_golden else ""
        cat_rows  += (f'<tr>'
                      f'<td><span class="cat-tag">{prefix}</span></td>'
                      f'<td class="num">{stats["count"]}</td>'
                      f'<td class="num">{stats["total_components"]}</td>'
                      f'<td class="num">{avg}</td>'
                      f'{iou_cell}'
                      f'<td>{pass_badge} {fail_badge}</td>'
                      f'</tr>\n')

    iou_th     = '<th>Avg IoU</th>' if has_golden else ''
    iou_chart  = ""
    iou_kpi    = ""
    pass_fail_chart = ""

    if has_golden:
        iou_kpi = f'<div class="kpi blue"><div class="val">{avg_iou if avg_iou else "—"}</div><div class="lbl">Avg IoU</div></div>'
        iou_chart = f"""
    <div class="card">
      <h3>IoU Score Distribution</h3>
      <canvas id="iouHist" height="200"></canvas>
    </div>"""
        pass_fail_chart = f"""
    <div class="card">
      <h3>PASS vs FAIL per Category</h3>
      <canvas id="passFail" height="200"></canvas>
    </div>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Detection Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root {{
  --bg:      #0a0f1e;
  --surface: #111827;
  --card:    #1a2236;
  --border:  #243048;
  --text:    #e2e8f0;
  --muted:   #64748b;
  --pass:    #22c55e;
  --fail:    #ef4444;
  --blue:    #3b82f6;
  --cyan:    #06b6d4;
  --purple:  #8b5cf6;
}}
* {{ box-sizing:border-box; margin:0; padding:0 }}
body {{ font-family:'Inter',sans-serif; background:var(--bg); color:var(--text); min-height:100vh }}

header {{
  background: linear-gradient(135deg, #0f1f3d 0%, #1a2a4d 50%, #162238 100%);
  border-bottom: 1px solid var(--border);
  padding: 0;
}}
.header-inner {{
  max-width:1400px; margin:0 auto; padding:20px 32px;
  display:flex; align-items:center; gap:20px;
}}
.header-icon {{
  width:48px; height:48px; border-radius:12px;
  background:linear-gradient(135deg,#3b82f6,#8b5cf6);
  display:flex; align-items:center; justify-content:center;
  font-size:22px; flex-shrink:0;
}}
.header-text h1 {{ font-size:20px; font-weight:700; letter-spacing:-0.3px }}
.header-text p  {{ font-size:12px; color:var(--muted); margin-top:3px; font-family:'JetBrains Mono',monospace }}
.mode-badge {{
  margin-left:auto; padding:6px 14px; border-radius:20px; font-size:11px;
  font-weight:600; font-family:'JetBrains Mono',monospace; letter-spacing:0.5px;
  background: {'#1a3a1a; color:#22c55e; border:1px solid #22c55e40' if has_golden else '#1a1a3a; color:#8b5cf6; border:1px solid #8b5cf640'};
}}

.container {{ max-width:1400px; margin:0 auto; padding:24px 32px }}

/* KPI */
.kpi-row {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(140px,1fr)); gap:14px; margin-bottom:24px }}
.kpi {{ background:var(--card); border:1px solid var(--border); border-radius:12px; padding:18px 14px; text-align:center; position:relative; overflow:hidden }}
.kpi::before {{ content:''; position:absolute; top:0; left:0; right:0; height:3px }}
.kpi.red::before   {{ background:#ef4444 }}
.kpi.green::before {{ background:#22c55e }}
.kpi.blue::before  {{ background:#3b82f6 }}
.kpi.purple::before {{ background:#8b5cf6 }}
.kpi.yellow::before {{ background:#f59e0b }}
.kpi .val {{ font-size:28px; font-weight:700; font-family:'JetBrains Mono',monospace }}
.kpi.red .val    {{ color:#ef4444 }}
.kpi.green .val  {{ color:#22c55e }}
.kpi.blue .val   {{ color:#3b82f6 }}
.kpi.purple .val {{ color:#8b5cf6 }}
.kpi.yellow .val {{ color:#f59e0b }}
.kpi .lbl {{ font-size:10px; color:var(--muted); text-transform:uppercase; letter-spacing:1px; margin-top:4px; font-weight:500 }}

/* Cards */
.row2 {{ display:grid; grid-template-columns:1fr 1fr; gap:18px; margin-bottom:18px }}
.row3 {{ display:grid; grid-template-columns:1fr 1fr 1fr; gap:18px; margin-bottom:18px }}
.card {{ background:var(--card); border:1px solid var(--border); border-radius:12px; padding:20px; margin-bottom:18px }}
.card h3 {{ font-size:11px; font-weight:600; color:var(--muted); text-transform:uppercase; letter-spacing:1.5px; margin-bottom:14px }}

/* Tables */
.tbl-wrap {{ overflow-x:auto }}
table {{ width:100%; border-collapse:collapse; font-size:12px }}
th {{ background:#0d1526; color:var(--muted); font-size:10px; font-weight:600; text-transform:uppercase; letter-spacing:1px; padding:9px 12px; text-align:left; border-bottom:1px solid var(--border); white-space:nowrap }}
td {{ padding:8px 12px; border-bottom:1px solid #1a2030; color:#cbd5e1 }}
tr:last-child td {{ border-bottom:none }}
tr:hover td {{ background:#1f2d44 }}
td.num {{ font-family:'JetBrains Mono',monospace; font-size:11px; color:#94a3b8 }}

/* Badges */
.badge {{ display:inline-block; padding:3px 8px; border-radius:6px; font-size:10px; font-weight:600; font-family:'JetBrains Mono',monospace; margin:1px }}
.badge.pass {{ background:#052e16; color:#22c55e; border:1px solid #22c55e30 }}
.badge.fail {{ background:#450a0a; color:#ef4444; border:1px solid #ef444430 }}
.cat-tag {{ display:inline-block; padding:3px 9px; border-radius:6px; font-size:10px; font-weight:600; background:#162038; color:#60a5fa; border:1px solid #3b82f630 }}
.dot {{ display:inline-block; width:8px; height:8px; border-radius:50%; margin-right:8px; vertical-align:middle }}
.pct {{ font-family:'JetBrains Mono',monospace; font-size:10px; color:var(--muted) }}
.bar-bg {{ height:4px; background:var(--border); border-radius:2px; margin-top:4px }}
.bar-fg {{ height:100%; border-radius:2px }}

/* Chart.js defaults */
@media(max-width:768px) {{
  .row2,.row3 {{ grid-template-columns:1fr }}
  .container,.header-inner {{ padding:16px }}
  .kpi .val {{ font-size:22px }}
}}
</style>
</head>
<body>
<header>
  <div class="header-inner">
    <div class="header-icon">🔬</div>
    <div class="header-text">
      <h1>Component Detection Dashboard</h1>
      <p>{ts_label} &nbsp;·&nbsp; {total_img} images &nbsp;·&nbsp; {'IoU threshold: ' + str(records[0]['iou_threshold']) if has_golden else 'Detection-only mode'}</p>
    </div>
    <div class="mode-badge">{'⚡ GOLDEN COMPARISON' if has_golden else '📦 DETECTION ONLY'}</div>
  </div>
</header>
<div class="container">

  <!-- KPIs -->
  <div class="kpi-row">
    <div class="kpi blue"><div class="val">{total_img}</div><div class="lbl">Total Images</div></div>
    {'<div class="kpi green"><div class="val">' + str(pass_cnt) + '</div><div class="lbl">PASS</div></div>' if has_golden else ''}
    {'<div class="kpi red"><div class="val">' + str(fail_cnt) + '</div><div class="lbl">FAIL</div></div>' if has_golden else ''}
    {'<div class="kpi green"><div class="val">' + str(pass_rate) + '%</div><div class="lbl">Pass Rate</div></div>' if has_golden else ''}
    {iou_kpi}
    <div class="kpi purple"><div class="val">{len(sorted_classes)}</div><div class="lbl">Classes</div></div>
    <div class="kpi yellow"><div class="val">{total_comp}</div><div class="lbl">Components</div></div>
    <div class="kpi blue"><div class="val">{avg_comp}</div><div class="lbl">Avg/Image</div></div>
  </div>

  <!-- Charts row 1 -->
  <div class="{'row3' if has_golden else 'row2'}">
    <div class="card"><h3>Components by Class</h3><canvas id="classBar" height="220"></canvas></div>
    <div class="card"><h3>Images by Category</h3><canvas id="catPie" height="220"></canvas></div>
    {pass_fail_chart if has_golden else ''}
  </div>

  <!-- Charts row 2 -->
  <div class="row2">
    <div class="card"><h3>Components per Category</h3><canvas id="catComp" height="200"></canvas></div>
    {iou_chart if has_golden else '<div class="card"><h3>Class Frequency (# images)</h3><canvas id="clsFreq" height="200"></canvas></div>'}
  </div>

  <!-- Class breakdown table -->
  <div class="card">
    <h3>Component Class Breakdown</h3>
    <div class="tbl-wrap">
      <table>
        <thead><tr><th>Class</th><th>Total</th><th>Images</th><th>Avg/Image</th><th>Share</th></tr></thead>
        <tbody>{cls_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- Category table -->
  <div class="card">
    <h3>Category Summary</h3>
    <div class="tbl-wrap">
      <table>
        <thead><tr><th>Category</th><th>Images</th><th>Components</th><th>Avg/Image</th>{iou_th}<th>Results</th></tr></thead>
        <tbody>{cat_rows}</tbody>
      </table>
    </div>
  </div>

</div>
<script>
const P  = ["#3b82f6","#06b6d4","#8b5cf6","#ec4899","#f97316","#10b981","#f59e0b","#60a5fa","#ef4444","#6366f1","#22c55e","#84cc16"];
const al = (cols, a) => cols.map(c => c + (a||'99'));
const g  = {{x:{{ticks:{{color:"#64748b",font:{{size:10}}}},grid:{{color:"#1a2540"}}}},y:{{ticks:{{color:"#64748b",font:{{size:10}}}},grid:{{color:"#243048"}}}}}};
const lg = {{labels:{{color:"#94a3b8",font:{{size:11}}}}}};

const clsL = {cls_labels}, clsT = {cls_totals}, clsI = {cls_imgs};
const catL = {cat_labels}, catC = {cat_counts}, catM = {cat_comps};
const catP = {json.dumps(cat_pass)}, catF = {json.dumps(cat_fail)};

new Chart(document.getElementById("classBar"),{{
  type:"bar",
  data:{{labels:clsL,datasets:[{{label:"Total",data:clsT,backgroundColor:al(P),borderColor:P,borderWidth:1,borderRadius:4}}]}},
  options:{{responsive:true,plugins:{{legend:{{display:false}}}},scales:g}}
}});

new Chart(document.getElementById("catPie"),{{
  type:"doughnut",
  data:{{labels:catL,datasets:[{{data:catC,backgroundColor:al(P.slice(0,catL.length),'cc'),borderColor:P,borderWidth:2}}]}},
  options:{{responsive:true,plugins:{{legend:lg}}}}
}});

new Chart(document.getElementById("catComp"),{{
  type:"bar",
  data:{{labels:catL,datasets:[{{label:"Components",data:catM,backgroundColor:al(P.slice(0,catL.length),'aa'),borderColor:P,borderWidth:1,borderRadius:4}}]}},
  options:{{responsive:true,plugins:{{legend:{{display:false}}}},scales:g}}
}});

{'new Chart(document.getElementById("passFail"),{type:"bar",data:{labels:catL,datasets:[{label:"PASS",data:catP,backgroundColor:"#22c55e55",borderColor:"#22c55e",borderWidth:1,borderRadius:3},{label:"FAIL",data:catF,backgroundColor:"#ef444455",borderColor:"#ef4444",borderWidth:1,borderRadius:3}]},options:{responsive:true,plugins:{legend:lg},scales:g}});' if has_golden else ''}

{'new Chart(document.getElementById("iouHist"),{type:"bar",data:{labels:' + json.dumps(iou_hist_labels) + ',datasets:[{label:"Images",data:' + json.dumps(iou_hist_vals) + ',backgroundColor:"#3b82f666",borderColor:"#3b82f6",borderWidth:1,borderRadius:3}]},options:{responsive:true,plugins:{legend:{display:false}},scales:g}});' if has_golden and iou_values else ''}

{'new Chart(document.getElementById("clsFreq"),{type:"bar",data:{labels:clsL,datasets:[{label:"Images",data:clsI,backgroundColor:"#8b5cf666",borderColor:"#8b5cf6",borderWidth:1,borderRadius:3}]},options:{responsive:true,indexAxis:"y",plugins:{legend:{display:false}},scales:g}});' if not has_golden else ''}
</script>
</body>
</html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✅ HTML  → {out_path}")

def build_pdf(records, sorted_classes, class_img_count, prefix_stats,
              status_counts, iou_values, has_golden, out_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, PageBreak)
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie

    W, H  = A4
    DKBL  = colors.HexColor("#1A3A5C")
    MDBL  = colors.HexColor("#2E75B6")
    LTBL  = colors.HexColor("#D6E8F7")
    ALTBL = colors.HexColor("#EBF3FB")
    PASS  = colors.HexColor("#1A7A3C")
    PASS_BG = colors.HexColor("#D8F5E4")
    FAIL  = colors.HexColor("#C00000")
    FAIL_BG = colors.HexColor("#FDDEDE")
    NA    = colors.HexColor("#888888")
    WHITE = colors.white
    GRAY  = colors.HexColor("#64748b")
    PAL   = [colors.HexColor(h) for h in
             ["#3b82f6","#06b6d4","#8b5cf6","#ec4899","#f97316",
              "#10b981","#f59e0b","#60a5fa","#ef4444","#6366f1","#22c55e"]]

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=1.4*cm, rightMargin=1.4*cm,
                            topMargin=1.4*cm, bottomMargin=1.4*cm)

    body_style = ParagraphStyle("body", fontSize=8, fontName="Helvetica",
                                textColor=colors.HexColor("#1e293b"))

    total_img  = len(records)
    fail_cnt   = status_counts.get("FAIL", 0)
    pass_cnt   = status_counts.get("PASS", 0)
    pass_rate  = round(pass_cnt/total_img*100, 1) if total_img else 0
    total_comp = sum(r["total_components"] for r in records)
    avg_iou    = round(sum(iou_values)/len(iou_values), 3) if iou_values else None
    ts_label   = str(records[0]["timestamp"]) if records else ""
    cat_data   = sorted(prefix_stats.items())
    PW         = W - 2.8*cm  # printable width

    def sec_bar(text, sub=""):
        d = Drawing(PW, 24)
        d.add(Rect(0, 0, PW, 24, fillColor=MDBL, strokeColor=None))
        d.add(String(8, 7, text, fontSize=10, fillColor=WHITE, fontName="Helvetica-Bold"))
        if sub:
            d.add(String(PW - 6, 7, sub, fontSize=8, fillColor=colors.HexColor("#93c5fd"),
                         fontName="Helvetica", textAnchor="end"))
        return d

    def mk_table(hdr, rows, cws, status_col=None):
        ts = TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), MDBL),
            ("TEXTCOLOR",   (0,0), (-1,0), WHITE),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [WHITE, ALTBL]),
            ("GRID",        (0,0), (-1,-1), 0.25, colors.HexColor("#c8d8ec")),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
        ])
        t = Table([hdr]+rows, colWidths=cws, repeatRows=1)
        t.setStyle(ts)
        if status_col is not None:
            for ri, row in enumerate(rows, 1):
                s = str(row[status_col]).upper()
                if s == "PASS":
                    t.setStyle(TableStyle([
                        ("BACKGROUND",(status_col,ri),(status_col,ri), PASS_BG),
                        ("TEXTCOLOR", (status_col,ri),(status_col,ri), PASS),
                        ("FONTNAME",  (status_col,ri),(status_col,ri), "Helvetica-Bold"),
                    ]))
                elif s == "FAIL":
                    t.setStyle(TableStyle([
                        ("BACKGROUND",(status_col,ri),(status_col,ri), FAIL_BG),
                        ("TEXTCOLOR", (status_col,ri),(status_col,ri), FAIL),
                        ("FONTNAME",  (status_col,ri),(status_col,ri), "Helvetica-Bold"),
                    ]))
        return t

    story = []

    cov = Drawing(PW, 70)
    cov.add(Rect(0, 0, PW, 70, fillColor=DKBL, strokeColor=None, rx=4, ry=4))
    mode_txt = "GOLDEN COMPARISON MODE" if has_golden else "DETECTION-ONLY MODE"
    cov.add(String(14, 44, "Component Detection Report", fontSize=20,
                   fillColor=WHITE, fontName="Helvetica-Bold"))
    cov.add(String(14, 26, f"{ts_label}   ·   {total_img} images   ·   {mode_txt}",
                   fontSize=9, fillColor=colors.HexColor("#93c5fd"), fontName="Helvetica"))
    story.append(cov)
    story.append(Spacer(1, 14))

    story.append(sec_bar("Overview"))
    story.append(Spacer(1, 8))

    kpi_items = [("Total Images", str(total_img), DKBL)]
    if has_golden:
        kpi_items += [
            ("PASS",      str(pass_cnt),      PASS),
            ("FAIL",      str(fail_cnt),      FAIL),
            ("Pass Rate", f"{pass_rate}%",    MDBL),
        ]
        if avg_iou: kpi_items.append(("Avg IoU", str(avg_iou), colors.HexColor("#0e7490")))
    kpi_items += [
        ("Classes",    str(len(sorted_classes)), colors.HexColor("#6d28d9")),
        ("Components", str(total_comp),           colors.HexColor("#b45309")),
    ]

    cols_n = min(4, len(kpi_items))
    rows_n = -(-len(kpi_items) // cols_n)
    cw = PW / cols_n
    ch = 46
    kd = Drawing(PW, rows_n * (ch + 4))
    for idx, (lbl, val, col) in enumerate(kpi_items):
        ci = idx % cols_n
        ri = idx // cols_n
        x  = ci * cw
        y  = (rows_n - 1 - ri) * (ch + 4)
        kd.add(Rect(x+1, y+1, cw-3, ch-2, fillColor=LTBL,
                    strokeColor=MDBL, strokeWidth=0.4, rx=3, ry=3))
        kd.add(String(x+cw/2, y+ch/2+3, val, fontSize=17, fillColor=col,
                       fontName="Helvetica-Bold", textAnchor="middle"))
        kd.add(String(x+cw/2, y+6, lbl, fontSize=7, fillColor=GRAY,
                       fontName="Helvetica", textAnchor="middle"))
    story.append(kd)
    story.append(Spacer(1, 16))

    cls_names = [c for c, _ in sorted_classes]
    cls_vals  = [t for _, t in sorted_classes]
    if cls_vals:
        story.append(sec_bar("Total Detected Components by Class"))
        story.append(Spacer(1, 8))
        bc = VerticalBarChart()
        bc.x, bc.y = 55, 20
        bc.width   = PW - 70
        bc.height  = 130
        bc.data    = [cls_vals]
        bc.categoryAxis.categoryNames = cls_names
        bc.categoryAxis.labels.angle  = 28
        bc.categoryAxis.labels.fontSize = 7
        bc.categoryAxis.labels.dy = -10
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 7
        bc.bars[0].fillColor = MDBL
        d1 = Drawing(PW, 170)
        d1.add(bc)
        story.append(d1)
        story.append(Spacer(1, 12))

    story.append(sec_bar("Image Distribution by Category"))
    story.append(Spacer(1, 8))
    pie_size   = 150
    total_cats = sum(s["count"] for _, s in cat_data) or 1
    pie_h      = pie_size + 30
    pie_d      = Drawing(PW, pie_h)
    pie        = Pie()
    pie.x      = 15
    pie.y      = 10
    pie.width  = pie.height = pie_size
    pie.data   = [s["count"] for _, s in cat_data]
    pie.labels = [f"{lbl}\n{round(s['count']/total_cats*100,1)}%" for lbl, s in cat_data]
    pie.sideLabels       = True
    pie.sideLabelsOffset = 0.08
    pie.simpleLabels     = False
    for i in range(len(cat_data)):
        pie.slices[i].fillColor   = PAL[i % len(PAL)]
        pie.slices[i].strokeWidth = 1
        pie.slices[i].strokeColor = WHITE
        pie.slices[i].fontSize    = 7
        pie.slices[i].fontName    = "Helvetica-Bold"
    pie_d.add(pie)
    story.append(pie_d)
    story.append(Spacer(1, 12))

    if has_golden:
        story.append(sec_bar("Pass / Fail Summary by Category",
                              f"IoU threshold: {records[0]['iou_threshold']}"))
        story.append(Spacer(1, 6))
        hdr = ["Category", "Images", "PASS", "FAIL", "Pass Rate", "Avg IoU"]
        rows = []
        for prefix, stats in cat_data:
            rate = f"{round(stats['pass']/stats['count']*100,1)}%" if stats['count'] else "—"
            avg_i = f"{round(stats['iou_sum']/stats['iou_count'],3):.3f}" if stats['iou_count'] else "—"
            rows.append([prefix, stats["count"], stats["pass"], stats["fail"], rate, avg_i])
        cws = [PW*f for f in [0.28, 0.12, 0.12, 0.12, 0.18, 0.18]]
        t = mk_table(hdr, rows, cws)
        for ri, row in enumerate(rows, 1):
            t.setStyle(TableStyle([
                ("TEXTCOLOR", (2,ri),(2,ri), PASS),
                ("FONTNAME",  (2,ri),(2,ri), "Helvetica-Bold"),
                ("TEXTCOLOR", (3,ri),(3,ri), FAIL),
                ("FONTNAME",  (3,ri),(3,ri), "Helvetica-Bold"),
            ]))
        story.append(t)
        story.append(Spacer(1, 12))

    story.append(sec_bar("Component Class Breakdown"))
    story.append(Spacer(1, 6))
    grand  = sum(cls_vals) or 1
    c_rows = []
    for cls, total in sorted_classes:
        img_c = class_img_count[cls]
        avg   = round(total/img_c, 2) if img_c else 0
        pct   = f"{round(total/grand*100,1)}%"
        c_rows.append([cls, total, img_c, avg, pct])
    cws2 = [PW*f for f in [0.36, 0.16, 0.18, 0.15, 0.15]]
    story.append(mk_table(["Class","Total","Images","Avg/Image","Share"], c_rows, cws2))
    story.append(Spacer(1, 12))

    story.append(PageBreak())
    story.append(sec_bar(f"Raw Detection Data — first 60 rows of {len(records)}"))
    story.append(Spacer(1, 6))

    if has_golden:
        raw_hdr  = ["Image Name","Status","Matched","Avg IoU","Category","Mode"]
        raw_rows = []
        for r in records[:60]:
            matched_s = f"{r['matched']}/{r['total']}" if r['matched'] is not None else "—"
            iou_s     = f"{r['avg_iou']:.3f}" if r['avg_iou'] is not None else "—"
            raw_rows.append([r["image_name"], r["status"], matched_s, iou_s,
                             r["prefix"], r["golden_mode"]])
        cws3 = [PW*f for f in [0.30, 0.10, 0.13, 0.12, 0.20, 0.15]]
        story.append(mk_table(raw_hdr, raw_rows, cws3, status_col=1))
    else:
        raw_hdr  = ["Image Name","Category","Detected Classes","Components"]
        raw_rows = []
        for r in records[:60]:
            dc = r["detected_classes"][:60] + ("…" if len(r["detected_classes"])>60 else "")
            raw_rows.append([r["image_name"], r["prefix"], dc, r["total_components"]])
        cws3 = [PW*f for f in [0.22, 0.14, 0.52, 0.12]]
        story.append(mk_table(raw_hdr, raw_rows, cws3))

    if len(records) > 60:
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"… and {len(records)-60} more rows in the Excel report.",
                                body_style))

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(GRAY)
        canvas.drawString(1.4*cm, 0.7*cm, "Component Detection Report — Auto-generated")
        canvas.drawRightString(W-1.4*cm, 0.7*cm, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    print(f"  ✅ PDF   → {out_path}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python detection_report_generator.py <input.csv or input.xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"❌ File not found: {path}")
        sys.exit(1)

    base     = os.path.splitext(path)[0]
    xlsx_out = base + "_report.xlsx"
    html_out = base + "_dashboard.html"
    pdf_out  = base + "_dashboard.pdf"

    print(f"\n📂 Reading: {path}")
    records, has_golden = load_data(path)
    fmt = "Format A (golden comparison)" if has_golden else "Format B (detection only)"
    print(f"   {len(records)} rows — {fmt}")

    STATUS_ORDER = {"FAIL": 0, "PASS": 1}
    records.sort(key=lambda r: STATUS_ORDER.get(r["status"].upper(), 2))

    sc, ci, ps, st, iou_vals = aggregate(records)

    print("\n📊 Excel…")
    build_excel(records, sc, ci, ps, st, iou_vals, has_golden, xlsx_out)
    print("🌐 HTML…")
    build_html(records, sc, ci, ps, st, iou_vals, has_golden, html_out)
    print("📄 PDF…")
    build_pdf(records, sc, ci, ps, st, iou_vals, has_golden, pdf_out)

    print(f"\n✨ Done!\n   {xlsx_out}\n   {html_out}\n   {pdf_out}\n")


if __name__ == "__main__":
    main()
